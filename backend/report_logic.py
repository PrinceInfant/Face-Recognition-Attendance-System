import pandas as pd
from sqlalchemy.orm import Session
import models, salary_logic
import datetime
import io
from openpyxl.styles import Font, Alignment, PatternFill

def generate_attendance_report(db: Session, employee_id: int = None, date: str = None):
    # If no date provided, default to today
    if not date:
        date = datetime.date.today().strftime("%Y-%m-%d")

    # 1. Setup Date Range
    try:
        target_date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        target_date = datetime.date.today()
        
    start_date = datetime.date(target_date.year, target_date.month, 1)

    # 2. Fetch Data
    employees = db.query(models.Employee).order_by(models.Employee.emp_id).all()
    if employee_id:
        employees = [e for e in employees if e.id == employee_id]
        
    att_records = db.query(models.Attendance).filter(
        models.Attendance.date >= start_date,
        models.Attendance.date <= target_date
    ).all()
    
    # Map: (emp_id, date) -> record
    att_map = {(a.employee_id, a.date): a for a in att_records}
    
    # Calculate Total Working Days in the Month (Excluding Sundays)
    next_month = target_date.replace(day=28) + datetime.timedelta(days=4)
    last_day_of_month = next_month - datetime.timedelta(days=next_month.day)
    
    total_working_days = 0
    d = datetime.date(target_date.year, target_date.month, 1)
    while d <= last_day_of_month:
        if d.weekday() != 6: # Not Sunday
            total_working_days += 1
        d += datetime.timedelta(days=1)
    
    # 3. Build Pivot Data
    pivot_data = []
    
    for emp in employees:
        row = {
            "ID": emp.emp_id,
            "Name": emp.name,
            "Dept": emp.department
        }
        
        present_count = 0
        absent_count = 0
        late_count = 0
        ot_count = 0
        half_day_count = 0
        
        current = start_date
        while current <= target_date:
            date_col = current.strftime("%d-%m-%Y") # Full Date Format
            is_sunday = (current.weekday() == 6)
            
            val = ""
            att = att_map.get((emp.id, current))
            
            if att:
                # Determine Status Code
                code = "P"
                
                if att.entry_time and att.entry_time.time() > datetime.time(9, 30):
                    late_count += 1
                
                # Half Day Check
                is_half_day = False
                if att.status == "Half Day":
                    is_half_day = True
                elif att.exit_time:
                    # Calculate standard hours (capped at 17:00)
                    shift_end = datetime.datetime.combine(current, datetime.time(17, 0))
                    standard_end = min(att.exit_time, shift_end)
                    duration = (standard_end - att.entry_time).total_seconds() / 3600
                    
                    if duration < 5.0:
                        is_half_day = True
                
                if is_half_day:
                    code = "P"
                    half_day_count += 1
                
                present_count += 1
                
                # Check Overtime (Exit after 17:00)
                if att.exit_time:
                     shift_end = datetime.datetime.combine(current, datetime.time(17, 0))
                     if att.exit_time > shift_end:
                         ot_count += 1
                
                # Just Code
                val = code
                
            elif is_sunday:
                val = "H"
            else:
                val = "A"
                absent_count += 1
            
            row[date_col] = val
            current += datetime.timedelta(days=1)
            
        row["Total Working Days"] = total_working_days
        row["Present"] = present_count
        row["Absent"] = absent_count
        row["Late Entry"] = late_count
        row["Half Day"] = half_day_count
        row["Overtime"] = ot_count
        pivot_data.append(row)
        
    df = pd.DataFrame(pivot_data)
    
    # 4. Generate Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance Register')
        
        if not df.empty:
            worksheet = writer.sheets['Attendance Register']
            
            # Styles
            header_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            
            for idx, col in enumerate(df.columns):
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = header_font
                cell.alignment = center_align
                
                if col in ["Name", "Dept"]:
                    width = 15
                elif col in ["ID"]:
                    width = 10
                elif col == "Total Working Days":
                    width = 20
                elif col == "Overtime":
                    width = 10
                elif col in ["Present", "Absent"]:
                     width = 8
                elif col in ["Late Entry", "Half Day"]:
                     width = 12
                else:
                    width = 12 # Fit DD-MM-YYYY
                
                col_letter = chr(65 + idx) if idx < 26 else f"A{chr(65 + (idx-26))}"
                worksheet.column_dimensions[col_letter].width = width

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_align
                    val_str = str(cell.value)
                    
                    if val_str == "A":
                        cell.font = Font(color="FF0000", bold=True) # Red
                    elif val_str in ["P", "L"]:
                        cell.font = Font(color="008000", bold=True) # Green
                    elif val_str == "HD":
                        cell.font = Font(color="0000FF", bold=True) # Blue
                    elif val_str == "H":
                        cell.font = Font(color="FF9900", bold=True) # Orange

    return output.getvalue()

def generate_daily_log_report(db: Session, date: str = None):
    # Default to today if no date
    if not date:
        date = datetime.date.today().strftime("%Y-%m-%d")

    target_date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
    
    # Get all employees
    employees = db.query(models.Employee).order_by(models.Employee.emp_id).all()
    
    # Get logs for the specific date
    logs = db.query(models.Attendance).filter(models.Attendance.date == target_date).all()
    log_map = {log.employee_id: log for log in logs}
    
    data = []
    for emp in employees:
        log = log_map.get(emp.id)
        
        entry = "--:--"
        exit = "--:--"
        duration_str = "--"
        work_hours_str = "--"
        late_str = "--"
        ot_str = "--"
        half_day_str = "--"
        status = "Absent"
        
        if log:
            status = log.status
            
            # Determine Half Day Status
            if status == "Half Day":
                half_day_str = "Yes"
            elif log.entry_time and log.exit_time:
                 diff = log.exit_time - log.entry_time
                 if diff.total_seconds() < 5 * 3600:
                     half_day_str = "Yes"
                 else:
                     half_day_str = "No"
            else:
                half_day_str = "No"

            if log.entry_time:
                entry = log.entry_time.strftime("%H:%M:%S")
                # Late Entry Check
                if log.entry_time.time() > datetime.time(9, 30):
                    late_str = "Yes"
                else:
                    late_str = "No"
            
            if log.exit_time:
                exit = log.exit_time.strftime("%H:%M:%S")
                
                # Total Duration
                diff = log.exit_time - log.entry_time
                total_sec = int(diff.total_seconds())
                h = total_sec // 3600
                m = (total_sec % 3600) // 60
                duration_str = f"{h}h {m}m"
                
                # Standard Work Hours (Capped at 17:00)
                shift_end = datetime.datetime.combine(log.date, datetime.time(17, 0))
                standard_end = min(log.exit_time, shift_end)
                std_diff = standard_end - log.entry_time
                std_sec = max(0, int(std_diff.total_seconds()))
                sh = std_sec // 3600
                sm = (std_sec % 3600) // 60
                work_hours_str = f"{sh}h {sm}m"
                
                # OT Hours
                if log.exit_time > shift_end:
                    ot_diff = log.exit_time - shift_end
                    ot_sec = int(ot_diff.total_seconds())
                    oh = ot_sec // 3600
                    om = (ot_sec % 3600) // 60
                    ot_str = f"{oh}h {om}m"
                else:
                    ot_str = "--"
                    
            elif log.entry_time:
                duration_str = "Active"
                
        elif target_date.weekday() == 6:
            status = "Holiday"
            duration_str = "Sunday"
            
        data.append({
            "Employee ID": emp.emp_id,
            "Name": emp.name,
            "Date": date,
            "Check In": entry,
            "Check Out": exit,
            "Late Entry": late_str,
            "Half Day": half_day_str,
            "Work Hours": work_hours_str,
            "Overtime": ot_str,
            "Total Duration": duration_str,
            "Status": status
        })
        
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Daily Log')
        
        if not df.empty:
            worksheet = writer.sheets['Daily Log']
            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center')
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            for idx, col in enumerate(df.columns):
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = header_font
                cell.alignment = alignment
                cell.fill = fill
                worksheet.column_dimensions[chr(65 + idx)].width = 20
                
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = alignment
                    if cell.value == "Absent": cell.font = Font(color="FF0000")
                    elif cell.value == "Present": cell.font = Font(color="008000")
                    
    return output.getvalue()

def generate_salary_report(db: Session, month: int, year: int):
    employees = db.query(models.Employee).order_by(models.Employee.emp_id.asc()).all()
    
    data = []
    for emp in employees:
        salary_info = salary_logic.calculate_monthly_salary(db, emp.id, month, year)
        if salary_info:
            data.append({
                "Employee ID": emp.emp_id,
                "Name": emp.name,
                "Department": emp.department,
                "Payroll Period": f"{month}/{year}",
                "Present Days": salary_info["present_days"],
                "Absent Days": salary_info["absent_days"],
                "Base Monthly Salary": emp.monthly_salary,
                "Pro-rata Salary": salary_info["calculated_salary"],
                "OT / Bonus": salary_info["ot_amount"],
                "Gross Payout": salary_info["final_salary"]
            })
            
    df = pd.DataFrame(data)
    
    # Calculate Summary Data
    total_disbursement = df["Gross Payout"].sum() if not df.empty else 0
    
    # Department-wise Breakdown
    if not df.empty:
        dept_summary = df.groupby("Department")["Gross Payout"].sum().reset_index()
        dept_summary.columns = ["Department Name", "Total Cost / Month"]
    else:
        dept_summary = pd.DataFrame(columns=["Department Name", "Total Cost / Month"])

    summary_data = [
        {"Metric": "Total Monthly Disbursement", "Value": total_disbursement},
        {"Metric": "Reporting Month", "Value": f"{month}/{year}"},
        {"Metric": "Total Employees Processed", "Value": len(df)}
    ]
    summary_df = pd.DataFrame(summary_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Detailed Employee Records
        df.to_excel(writer, index=False, sheet_name='Employee Payroll')
        
        # Sheet 2: Financial Summary
        summary_df.to_excel(writer, index=False, sheet_name='Financial Summary')
        dept_summary.to_excel(writer, index=False, sheet_name='Financial Summary', startrow=len(summary_df) + 3)
        
        # Formatting
        header_font = Font(bold=True)
        alignment = Alignment(horizontal='center')
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for sheet_name in ['Employee Payroll', 'Financial Summary']:
            worksheet = writer.sheets[sheet_name]
            current_df = df if sheet_name == 'Employee Payroll' else summary_df
            
            # Style Headers
            for idx, col in enumerate(current_df.columns):
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = header_font
                cell.alignment = alignment
                cell.fill = header_fill
                worksheet.column_dimensions[chr(65 + idx)].width = 25
            
            # Style Body
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = alignment

    return output.getvalue()

def generate_department_salary_report(db: Session, month: int, year: int):
    employees = db.query(models.Employee).all()
    
    dept_data = []
    for emp in employees:
        salary_info = salary_logic.calculate_monthly_salary(db, emp.id, month, year)
        if salary_info:
            dept_data.append({
                "Department": emp.department,
                "Salary": salary_info["final_salary"]
            })
            
    if not dept_data:
        df = pd.DataFrame(columns=["Department Name", "Total Salary Cost", "Payroll Period"])
    else:
        temp_df = pd.DataFrame(dept_data)
        df = temp_df.groupby("Department")["Salary"].sum().reset_index()
        df.columns = ["Department Name", "Total Salary Cost"]
        df["Payroll Period"] = f"{month}/{year}"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dept Summary')
        
        if not df.empty:
            worksheet = writer.sheets['Dept Summary']
            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center')
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            for idx, col in enumerate(df.columns):
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = header_font
                cell.alignment = alignment
                cell.fill = header_fill
                worksheet.column_dimensions[chr(65 + idx)].width = 25
                
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = alignment
            
    return output.getvalue()

def generate_employee_directory_report(db: Session):
    employees = db.query(models.Employee).all()
    
    data = []
    for emp in employees:
        data.append({
            "Employee ID": emp.emp_id,
            "Name": emp.name,
            "Sector": emp.department,
            "Base Salary": emp.monthly_salary,
            "OT Rule": f"{emp.ot_rule}x",
            "Registration Date": datetime.date.today().strftime("%Y-%m-%d")
        })
            
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Workforce Directory')
        
        if not df.empty:
            worksheet = writer.sheets['Workforce Directory']
            header_font = Font(bold=True)
            alignment = Alignment(horizontal='center')
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Style Headers
            for idx, col in enumerate(df.columns):
                cell = worksheet.cell(row=1, column=idx+1)
                cell.font = header_font
                cell.alignment = alignment
                cell.fill = header_fill
                worksheet.column_dimensions[chr(65 + idx)].width = 25
                
            # Style Body
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = alignment
            
    return output.getvalue()
